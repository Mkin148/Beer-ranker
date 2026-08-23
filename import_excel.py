"""
Seed the Meet and Drink database from the original Excel spreadsheet.

Excel column -> app field:
    A Beer name          -> name
    B Style/ABV text     -> style   (ABV parsed out of it)
    C Look               -> look
    D Drinkability       -> drinkability
    E Taste              -> taste
    F Smell              -> smell
    G "Description"*      -> packaging  (*header mislabelled a numeric 5th score)
    J Notes              -> description

Usage:
    python import_excel.py Beer_tasting_.xlsx
Rows with no name or no scores are skipped.
"""

import sys
from openpyxl import load_workbook

import app  # reuses the DB layer + schema

# Excel column index (0-based) -> app score key
EXCEL_COL_TO_KEY = {2: "look", 3: "drinkability", 4: "taste", 5: "smell", 6: "packaging"}


def main(path):
    app.init_db()
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    added, skipped = 0, 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        name = (row[0] or "").strip() if row and row[0] else ""
        raw = {idx: row[idx] if len(row) > idx else None for idx in EXCEL_COL_TO_KEY}
        if not name or not any(isinstance(v, (int, float)) for v in raw.values()):
            skipped += 1
            continue

        style = (row[1] or "").strip() if len(row) > 1 and row[1] else ""
        scores = {
            key: float(raw[idx]) if isinstance(raw[idx], (int, float)) else 0.0
            for idx, key in EXCEL_COL_TO_KEY.items()
        }
        description = row[9].strip() if len(row) > 9 and isinstance(row[9], str) else ""

        app.add_beer(name, style, app.parse_abv(style), scores, description, None, None)
        added += 1

    print(f"Imported {added} beers, skipped {skipped} rows.")


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "Beer_tasting_.xlsx"
    main(src)
